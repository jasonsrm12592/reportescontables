import streamlit as st
import xmlrpc.client
import pandas as pd
import io
import datetime
from openpyxl import styles 

# --- CONFIGURACIÓN GLOBAL ---
st.set_page_config(page_title="Reportes Contables Odoo", layout="wide", page_icon="📊")

# ==========================================
# 1. BACKEND: CONEXIÓN Y PROCESAMIENTO
# ==========================================

def get_odoo_connection():
    try:
        url = st.secrets["odoo"]["url"]
        db = st.secrets["odoo"]["db"]
        username = st.secrets["odoo"]["username"]
        password = st.secrets["odoo"]["password"]
        
        common = xmlrpc.client.ServerProxy('{}/xmlrpc/2/common'.format(url))
        uid = common.authenticate(db, username, password, {})
        models = xmlrpc.client.ServerProxy('{}/xmlrpc/2/object'.format(url))
        return uid, models, db, password
    except Exception as e:
        st.error(f"Error de conexión: {e}")
        return None, None, None, None

def fetch_data(uid, models, db, password, cutoff_date):
    # 1. Traer líneas contables (SOLO COMPAÑÍA 1)
    # QUITAMOS 'x_studio_es_reintegro' de aquí porque suele estar en la cabecera, no en la línea.
    domain = [
        ('parent_state', '=', 'posted'),
        ('company_id', '=', 1),
        ('account_type', '=', 'liability_payable'),
        ('amount_residual', '!=', 0),
        ('move_id.move_type', 'in', ['in_invoice', 'in_refund']),
    ]
    
    fields = ['partner_id', 'date_maturity', 'date', 'ref', 
              'amount_residual', 'amount_residual_currency', 
              'currency_id', 'move_id'] # <--- Quitamos el campo de Studio de esta lista
    
    lines = models.execute_kw(db, uid, password, 'account.move.line', 'search_read', [domain], {'fields': fields})
    
    if not lines: return pd.DataFrame()

    df = pd.DataFrame(lines)
    
    # 2. CONSULTA ADICIONAL A LA CABECERA (account.move)
    # Aquí buscamos el campo de Reintegro y el Tipo de Documento
    move_ids_list = [m[0] for m in df['move_id'] if m]
    move_ids_unique = list(set(move_ids_list))
    
    type_map = {}
    reintegro_map = {}
    
    if move_ids_unique:
        # Pedimos los campos a la cabecera
        move_fields = ['move_type', 'x_studio_es_reintegro']
        moves_data = models.execute_kw(db, uid, password, 'account.move', 'search_read', 
                                       [[('id', 'in', move_ids_unique)]], 
                                       {'fields': move_fields})
        
        # Mapeamos los resultados
        for m in moves_data:
            type_map[m['id']] = m['move_type']
            # Ojo: x_studio_es_reintegro podría no venir si el campo no existe en algunas facturas viejas
            reintegro_map[m['id']] = m.get('x_studio_es_reintegro', False)

    # 3. Limpieza y Mapeo
    df['Proveedor'] = df['partner_id'].apply(lambda x: x[1] if x else 'Sin Proveedor')
    df['Partner_ID'] = df['partner_id'].apply(lambda x: x[0] if x else False)
    df['Moneda'] = df['currency_id'].apply(lambda x: x[1] if x else '')
    df['ref'] = df['ref'].apply(lambda x: x if x else '-')
    df['move_id_int'] = df['move_id'].apply(lambda x: x[0] if x else False)
    
    # APLICAMOS EL MAPEO DEL CAMPO STUDIO AL DATAFRAME
    df['x_studio_es_reintegro'] = df['move_id_int'].map(reintegro_map).fillna(False)
    
    # Corrección Fechas Vacías
    df['date_maturity'] = df.apply(lambda row: row['date'] if not row['date_maturity'] else row['date_maturity'], axis=1)
    df['date_maturity'] = pd.to_datetime(df['date_maturity'], errors='coerce')
    df = df.dropna(subset=['date_maturity'])

    # 4. Lógica de Montos (Moneda Original)
    def calcular_neto(row):
        tipo = type_map.get(row['move_id_int'], 'in_invoice')
        
        if row['amount_residual_currency'] and row['amount_residual_currency'] != 0:
            monto_base = row['amount_residual_currency']
        else:
            monto_base = row['amount_residual']
            
        monto_abs = abs(monto_base)
        
        if tipo == 'in_refund': 
            return -monto_abs
        return monto_abs

    df['amount_residual_neto'] = df.apply(calcular_neto, axis=1)

    # 5. Buckets
    fecha_corte_pd = pd.to_datetime(cutoff_date)
    df['dias_vencido'] = (fecha_corte_pd - df['date_maturity']).dt.days
    
    cols_bucket = ['En Fecha', '1-30', '31-60', '61-90', '+90']
    for col in cols_bucket: df[col] = 0.0

    def clasificar_monto(row):
        dias = row['dias_vencido']
        monto = row['amount_residual_neto']
        if dias <= 0: row['En Fecha'] = monto
        elif 1 <= dias <= 30: row['1-30'] = monto
        elif 31 <= dias <= 60: row['31-60'] = monto
        elif 61 <= dias <= 90: row['61-90'] = monto
        else: row['+90'] = monto
        return row

    df = df.apply(clasificar_monto, axis=1)
    df['date_maturity'] = df['date_maturity'].dt.date
    
    return df

def normalize_currency_code(text):
    if not text: return None
    t = str(text).lower().strip()
    if 'colon' in t or 'crc' in t: return 'CRC'
    if 'dolar' in t or 'dólar' in t or 'usd' in t: return 'USD'
    return None

def detect_currency_in_obs(obs_text):
    if not obs_text: return None
    t = str(obs_text).lower()
    if 'dolar' in t or 'dólar' in t or 'usd' in t: return 'USD'
    if 'colon' in t or 'crc' in t: return 'CRC'
    return None

def enrich_with_smart_banks_split(df, models, uid, db, password):
    if df.empty: return df
    
    partner_ids = [p for p in df['Partner_ID'].unique().tolist() if p]
    if not partner_ids: 
        df['Banco'] = ''
        df['Cuenta'] = ''
        df['Notas Banco'] = ''
        return df

    bank_domain = [
        ('partner_id', 'in', partner_ids),
        '|', ('company_id', '=', False), ('company_id', '=', 1)
    ]
    
    bank_fields = ['partner_id', 'bank_id', 'acc_number', 'x_studio_observacin', 'currency_id']
    banks_data = models.execute_kw(db, uid, password, 'res.partner.bank', 'search_read', [bank_domain], {'fields': bank_fields})
    
    banks_by_partner = {}
    for b in banks_data:
        p_id = b['partner_id'][0]
        
        banco_name = b['bank_id'][1] if b['bank_id'] else "Banco"
        cuenta_num = b['acc_number'] or ""
        obs_txt = b.get('x_studio_observacin') or ""
        moneda_obs = detect_currency_in_obs(obs_txt)
        moneda_oficial = normalize_currency_code(b['currency_id'][1] if b['currency_id'] else None)
        
        bank_obj = {
            'banco': banco_name,
            'cuenta': cuenta_num,
            'obs': obs_txt,
            'moneda_obs': moneda_obs,
            'moneda_oficial': moneda_oficial
        }
        
        if p_id not in banks_by_partner: banks_by_partner[p_id] = []
        banks_by_partner[p_id].append(bank_obj)

    def get_best_bank_columns(row):
        p_id = row['Partner_ID']
        moneda_factura = normalize_currency_code(row['Moneda'])
        default_res = pd.Series(['', '', ''], index=['Banco', 'Cuenta', 'Notas Banco'])
        
        if p_id not in banks_by_partner: return default_res
        mis_bancos = banks_by_partner[p_id]
        
        matches_obs = [b for b in mis_bancos if b['moneda_obs'] == moneda_factura]
        if matches_obs:
            best = matches_obs[0]
            return pd.Series([best['banco'], best['cuenta'], best['obs']], index=['Banco', 'Cuenta', 'Notas Banco'])

        matches_field = [b for b in mis_bancos if b['moneda_oficial'] == moneda_factura]
        if matches_field:
            best = matches_field[0]
            return pd.Series([best['banco'], best['cuenta'], best['obs']], index=['Banco', 'Cuenta', 'Notas Banco'])
        
        matches_any = [b for b in mis_bancos if b['moneda_obs'] is None and b['moneda_oficial'] is None]
        if matches_any:
            best = matches_any[0]
            return pd.Series([best['banco'], best['cuenta'], best['obs']], index=['Banco', 'Cuenta', 'Notas Banco'])
        
        if mis_bancos:
             best = mis_bancos[0]
             return pd.Series([best['banco'], best['cuenta'], best['obs']], index=['Banco', 'Cuenta', 'Notas Banco'])

        return default_res

    bank_cols = df.apply(get_best_bank_columns, axis=1)
    df = pd.concat([df, bank_cols], axis=1)
    
    return df

# ==========================================
# 2. GENERACIÓN DE EXCEL MULTI-HOJA
# ==========================================

def clasificar_factura(row):
    """Define a qué hoja va la factura según prioridad"""
    
    # Prioridad 1: Reintegros (Campo Studio activado)
    if row.get('x_studio_es_reintegro') is True:
        return 'Reintegros Cajas Chicas'
    
    # Prioridad 2: Versatec (Referencia contiene VF)
    ref = str(row.get('ref', '')).upper()
    if "VF" in ref:
        return 'Versatec'
    
    # Prioridad 3: Por Moneda
    moneda = normalize_currency_code(row.get('Moneda', ''))
    if moneda == 'USD':
        return 'Dólares'
    else:
        return 'Colones'

def generar_excel_agrupado(df):
    output = io.BytesIO()
    
    # 1. Aplicar clasificación
    df['Hoja_Destino'] = df.apply(clasificar_factura, axis=1)
    
    # Columnas a exportar
    cols_export = ['ref', 'date', 'date_maturity', 'dias_vencido', 'Moneda', 
                   'En Fecha', '1-30', '31-60', '61-90', '+90', 'amount_residual_neto', 
                   'Banco', 'Cuenta', 'Notas Banco'] 
    
    header_names = ['Referencia', 'Emisión', 'Vencimiento', 'Días Vencido', 'Moneda', 
                    'Por Vencer', '1-30 Días', '31-60 Días', '61-90 Días', '+90 Días', 'Total',
                    'Banco', 'Cuenta', 'Notas']

    # 2. Escribir Excel con múltiples hojas
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        
        # Obtenemos las categorías únicas para crear las hojas
        # Forzamos un orden específico si existen datos
        orden_hojas = ['Colones', 'Dólares', 'Reintegros Cajas Chicas', 'Versatec']
        hojas_existentes = df['Hoja_Destino'].unique().tolist()
        
        for nombre_hoja in orden_hojas:
            if nombre_hoja not in hojas_existentes:
                continue # Si no hay datos para esta hoja, la saltamos (o podrías crearla vacía)

            # Filtramos datos para esta hoja
            df_hoja = df[df['Hoja_Destino'] == nombre_hoja].copy()
            
            # Ordenamos
            df_hoja = df_hoja.sort_values(by=['Proveedor', 'dias_vencido'], ascending=[True, False])
            
            # Creamos la hoja
            workbook = writer.book
            worksheet = workbook.create_sheet(nombre_hoja)
            
            # Estilos
            bold_font = styles.Font(bold=True)
            white_font = styles.Font(bold=True, color="FFFFFF")
            header_fill = styles.PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            prov_fill = styles.PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
            
            # Encabezados
            for col_idx, val in enumerate(header_names, 1):
                cell = worksheet.cell(row=1, column=col_idx, value=val)
                cell.font = white_font
                cell.fill = header_fill
            
            current_row = 2
            proveedores = df_hoja['Proveedor'].unique()
            
            for prov in proveedores:
                cell_title = worksheet.cell(row=current_row, column=1, value=f"PROVEEDOR: {prov}")
                cell_title.font = bold_font
                cell_title.fill = prov_fill
                worksheet.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=len(header_names))
                current_row += 1
                
                sub_df = df_hoja[df_hoja['Proveedor'] == prov][cols_export]
                
                for _, row in sub_df.iterrows():
                    for col_idx, value in enumerate(row, 1):
                        cell = worksheet.cell(row=current_row, column=col_idx, value=value)
                        if 6 <= col_idx <= 11: 
                            cell.number_format = '#,##0.00'
                    current_row += 1
                current_row += 1

            # Ajuste de ancho
            for col in worksheet.columns:
                max_len = 0
                col_letter = col[0].column_letter 
                for cell in col:
                    try:
                        if len(str(cell.value)) > max_len: max_len = len(str(cell.value))
                    except: pass
                worksheet.column_dimensions[col_letter].width = min(max_len + 2, 40)
        
        # Eliminar la hoja "Sheet" que se crea por defecto si no se usó
        if "Sheet" in writer.book.sheetnames:
            del writer.book["Sheet"]

    return output.getvalue()

# ==========================================
# 3. VISTAS
# ==========================================

def vista_inicio():
    st.title("🏠 Portal Financiero")
    st.markdown("Bienvenido. Genera tus reportes desde el menú lateral.")

def vista_reporte():
    st.title("📊 Cuentas por pagar")
    st.divider()

    col1, col2 = st.columns([1, 3])
    with col1:
        st.subheader("Configuración")
        f_corte = st.date_input("Fecha de Corte", pd.to_datetime("today"))
        btn = st.button("Generar Reporte", type="primary")

    with col2:
        if btn:
            with st.spinner('Procesando...'):
                uid, models, db, pwd = get_odoo_connection()
                if uid:
                    df = fetch_data(uid, models, db, pwd, f_corte)
                    if not df.empty:
                        df = enrich_with_smart_banks_split(df, models, uid, db, pwd)
                        
                        # --- VISTA PREVIA ---
                        # Clasificamos también para la vista previa
                        df['Hoja_Destino'] = df.apply(clasificar_factura, axis=1)
                        
                        st.subheader("Resumen General")
                        
                        # Mostramos un contador por categoría
                        resumen = df['Hoja_Destino'].value_counts()
                        st.dataframe(resumen, use_container_width=True)
                        
                        cols_view = ['Proveedor', 'ref', 'date_maturity', 'dias_vencido', 
                                     'amount_residual_neto', 'Moneda', 'Hoja_Destino']
                        
                        df_display = df.sort_values(by='dias_vencido', ascending=False)[cols_view]

                        st.dataframe(
                            df_display.style.format({'amount_residual_neto': "{:,.2f}"})
                            .map(lambda x: 'color: #d9534f' if x > 0 else 'color: black', subset=['dias_vencido']),
                            use_container_width=True
                        )
                        
                        excel_data = generar_excel_agrupado(df)
                        st.download_button("📥 Descargar Excel Multi-Hoja", excel_data, f"Antiguedad_{f_corte}.xlsx", "application/vnd.ms-excel")
                    else:
                        st.warning("No hay datos.")

# ==========================================
# 4. REPORTE VENTAS RETAIL
# ==========================================

def fetch_retail_sales(uid, models, db, password, start_date, end_date):
    """
    Trae facturas y notas de crédito de clientes (out_invoice, out_refund)
    FILTRADO POR VENDEDORES ESPECÍFICOS RETAIL.
    Lista: ALEJANDRO HERNANDEZ , GREIVIN VASQUEZ , JACKSON ABARCA , JOHNSEN MONTERO , LEONARDO CORRALES , SEBASTIAN CARRILLO
    """
    
    # 1. Traer TODAS las facturas en el rango (Company 1, Posted)
    domain_moves = [
        ('state', '=', 'posted'),
        ('company_id', '=', 1),
        ('move_type', 'in', ['out_invoice', 'out_refund']),
        ('invoice_date', '>=', str(start_date)),
        ('invoice_date', '<=', str(end_date)),
    ]
    
    fields_moves = ['name', 'invoice_date', 'invoice_user_id', 'amount_untaxed_signed', 'move_type', 'partner_id']
    
    moves = models.execute_kw(db, uid, password, 'account.move', 'search_read', [domain_moves], {'fields': fields_moves})
    
    if not moves:
        return pd.DataFrame()
        
    df = pd.DataFrame(moves)
    
    # 2. Filtrar por Vendedor (Case Insensitive Match)
    # Lista de vendedores Retail
    vendedores_retail = [
        "ALEJANDRO HERNANDEZ", 
        "GREIVIN VASQUEZ", 
        "JACKSON ABARCA", 
        "JOHNSEN MONTERO", 
        "LEONARDO CORRALES", 
        "SEBASTIAN CARRILLO"
    ]
    
    # Normalizamos para comparar (mayúsculas)
    vendedores_retail_norm = [v.upper() for v in vendedores_retail]
    
    def es_vendedor_retail(user_tuple):
        if not user_tuple: return False
        # user_tuple es (id, "Nombre")
        nombre = str(user_tuple[1]).upper()
        # Verificamos si alguno de los nombres retail está contenido en el nombre del vendedor Odoo
        # Usamos coincidencia exacta o substring según preferencia. 
        # Dado que los nombres parecen completos, intentaremos coincidencia parcial fuerte o exacta.
        for v in vendedores_retail_norm:
            if v in nombre:
                return True
        return False

    df['Vendedor_Raw'] = df['invoice_user_id']
    df = df[df['Vendedor_Raw'].apply(es_vendedor_retail)].copy()
    
    if df.empty:
        return pd.DataFrame()
    
    # 3. Procesamiento
    df['Vendedor'] = df['invoice_user_id'].apply(lambda x: x[1] if x else 'Sin Vendedor')
    df['Cliente'] = df['partner_id'].apply(lambda x: x[1] if x else 'Sin Cliente')
    df['Fecha'] = pd.to_datetime(df['invoice_date'])
    df['Mes'] = df['Fecha'].dt.strftime('%Y-%m') # Agrupación Mensual
    
    # Renombrar para visualización
    df = df.rename(columns={
        'name': 'Número Factura',
        'amount_untaxed_signed': 'Monto (CRC) Antes Imp.'
    })
    
    return df

def vista_ventas_retail():
    st.title("🛍️ Ventas Netas Retail")
    st.markdown("Facturas y Notas de Crédito confirmadas **sin cuenta analítica**.")
    st.divider()

    col1, col2 = st.columns([1, 3])
    with col1:
        st.subheader("Filtros")
        
        # Default: Mes actual
        today = datetime.date.today()
        first_day = today.replace(day=1)
        
        f_inicio = st.date_input("Fecha Inicio", first_day)
        f_fin = st.date_input("Fecha Fin", today)
        
        btn = st.button("Generar Reporte Retail", type="primary")

    with col2:
        if btn:
            with st.spinner('Consultando Odoo...'):
                uid, models, db, pwd = get_odoo_connection()
                if uid:
                    df = fetch_retail_sales(uid, models, db, pwd, f_inicio, f_fin)
                    
                    if not df.empty:
                        # KPI Totals
                        total_ventas = df['Monto (CRC) Antes Imp.'].sum()
                        
                        st.metric(label="Total Ventas Netas (Sin Impuestos)", value=f"₡ {total_ventas:,.2f}")
                        
                        # Agrupado por Mes
                        st.subheader("Desglose Mensual")
                        df_grouped = df.groupby('Mes')['Monto (CRC) Antes Imp.'].sum().reset_index()
                        st.dataframe(df_grouped.style.format({'Monto (CRC) Antes Imp.': "₡ {:,.2f}"}), use_container_width=True)
                        
                        # Detalle
                        st.subheader("Detalle de Facturas")
                        cols_view = ['Fecha', 'Mes', 'Número Factura', 'Cliente', 'Vendedor', 'Monto (CRC) Antes Imp.']
                        st.dataframe(
                            df[cols_view].sort_values(by='Fecha', ascending=False)
                            .style.format({'Monto (CRC) Antes Imp.': "{:,.2f}"}),
                            use_container_width=True
                        )
                        
                        # Excel Download
                        output = io.BytesIO()
                        with pd.ExcelWriter(output, engine='openpyxl') as writer:
                            df[cols_view].to_excel(writer, index=False, sheet_name='Detalle Retail')
                            df_grouped.to_excel(writer, index=False, sheet_name='Resumen Mensual')
                            
                        st.download_button(
                            "📥 Descargar Excel Retail", 
                            output.getvalue(), 
                            f"Ventas_Retail_{f_inicio}_al_{f_fin}.xlsx", 
                            "application/vnd.ms-excel"
                        )
                        
                    else:
                        st.info("No se encontraron registros para los filtros seleccionados (o todas las facturas tienen cuenta analítica).")

# ==========================================
# 5. REPORTE WIP (PROYECTOS EN PROCESO)
# ==========================================

# ==========================================
# 5. REPORTE WIP (PROYECTOS EN PROCESO)
# ==========================================

def fetch_wip_data(uid, models, db, password, end_date):
    """
    Reporte para analizar la cuenta 0.11531 WIP(Proyecto en Proceso) [ID 503].
    Cruza con facturación real y facturación estimada (x_facturas.proyectos).
    REFACTOR: WIP calculado desde account.move.line usando analytic_distribution (JSON).
    """
    end_date_str = str(end_date)
    start_of_month = end_date.replace(day=1)
    
    # 1. SALDO WIP (GASTOS PENDIENTES) - DESDE APUNTES CONTABLES
    # Usamos account.move.line porque el usuario valida contra analytic_distribution
    # REQ: Solo apuntes NO conciliados
    domain_wip = [
        ('parent_state', '=', 'posted'),
        ('account_id', '=', 503),
        ('date', '<=', end_date_str),
        ('company_id', '=', 1),
        ('reconciled', '!=', True) # Solo partidas abiertas (False or None)
    ]
    
    fields_wip = ['debit', 'credit', 'analytic_distribution', 'name', 'move_id']
    
    try:
        # Traemos todos los apuntes de la cuenta 503
        lines_wip = models.execute_kw(db, uid, password, 'account.move.line', 'search_read', [domain_wip], {'fields': fields_wip})
        df_wip = pd.DataFrame(lines_wip)
    except Exception as e:
        st.error(f"Error fetching WIP (Move Lines): {e}")
        return pd.DataFrame()

    if df_wip.empty:
        return pd.DataFrame()
        
    # Formato esperado: {'3686': 100} donde key es el ID de la cuenta analítica
    
    def extract_analytic_id(dist):
        if not dist: return None
        # Si es un diccionario (lo normal en xmlrpc para Json cols)
        if isinstance(dist, dict) and dist:
            # Retornamos el primer ID encontrado. 
            # Si hay split (varios IDs), esto tomaría uno arbitrario, pero para WIP suele ser 1 a 1.
            return list(dist.keys())[0]
        # Si Odoo devuelve string (menos probable via xmlrpc pero posible)
        return None

    df_wip['Analytic_ID_Str'] = df_wip['analytic_distribution'].apply(extract_analytic_id)
    
    # Filtrar lineas que no tienen analítica (si las hay, quedarían como 'Sin Proyecto')
    # Convertir a entero
    df_wip['Project_ID'] = pd.to_numeric(df_wip['Analytic_ID_Str'], errors='coerce')
    
    # Calcular Balance (Debit - Credit)
    # En apuntes contables: Gasto aumenta al Debe (Debit).
    # Saldo WIP = Debit - Credit.
    df_wip['WIP_Balance'] = df_wip['debit'] - df_wip['credit']
    
    # Agrupar por Project ID
    wip_grouped = df_wip.groupby('Project_ID')['WIP_Balance'].sum().reset_index()
    wip_grouped = wip_grouped[wip_grouped['WIP_Balance'] != 0]
    
    project_ids = wip_grouped['Project_ID'].dropna().unique().tolist()
    
    if not project_ids:
        return pd.DataFrame()

    # Necesitamos los Nombres de los Proyectos (Analytic Accounts)
    # Hicimos groupby solo por ID, ahora buscamos los nombres
    # project_ids son IDs de analytic.account
    try:
        # Include archived accounts to resolve names for closed projects
        analytic_accounts = models.execute_kw(db, uid, password, 'account.analytic.account', 'search_read', 
                                              [[('id', 'in', project_ids)]], 
                                              {'fields': ['id', 'name', 'active'], 'context': {'active_test': False}})
        aa_map = {a['id']: a['name'] for a in analytic_accounts}
        # active=True -> Archived=False. active=False -> Archived=True
        aa_archived_map = {a['id']: not a['active'] for a in analytic_accounts}
    except Exception as e:
        aa_map = {}
        aa_archived_map = {}
        
    wip_grouped['Project_Name'] = wip_grouped['Project_ID'].apply(lambda x: aa_map.get(x, f"Proyecto {int(x)}"))
    wip_grouped['Is_Archived'] = wip_grouped['Project_ID'].apply(lambda x: aa_archived_map.get(x, False))

    # 2. FACTURACIÓN REAL (INGRESOS)
    # Mantenemos account.analytic.line para ingresos pues es más eficiente buscar por account_id (Analytic ID)
    # Buscamos ingresos asociados a estos proyectos.
    
    # Primero identificamos cuentas financieras de ingreso
    try:
        income_account_ids = models.execute_kw(db, uid, password, 'account.account', 'search', [[('account_type', 'in', ['income', 'income_other'])]])
    except:
        income_account_ids = []
        
    domain_income = [
        ('company_id', '=', 1),
        ('account_id', 'in', project_ids), # account_id en analytic_line ES la cuenta analítica
        ('date', '<=', end_date_str)
    ]
    
    if income_account_ids:
        domain_income.append(('general_account_id', 'in', income_account_ids))
    
    fields_income = ['account_id', 'amount', 'date']
    try:
        lines_income = models.execute_kw(db, uid, password, 'account.analytic.line', 'search_read', [domain_income], {'fields': fields_income})
        df_income = pd.DataFrame(lines_income)
    except:
        df_income = pd.DataFrame() # Fail gracefully
    
    # Link to Project
    income_map_current = {}
    income_map_prev = {}
    
    if not df_income.empty:
        # account_id in analytic line is relations (id, name)
        df_income['Project_ID'] = df_income['account_id'].apply(lambda x: x[0] if x else None)
        df_income['Net_Income'] = df_income['amount'] # Income is positive in analytic
        df_income['Date'] = pd.to_datetime(df_income['date']).dt.date
        
        # Split Current vs Previous
        current_mask = (df_income['Date'] >= start_of_month) & (df_income['Date'] <= end_date)
        prev_mask = (df_income['Date'] < start_of_month)
        
        income_map_current = df_income[current_mask].groupby('Project_ID')['Net_Income'].sum().to_dict()
        income_map_prev = df_income[prev_mask].groupby('Project_ID')['Net_Income'].sum().to_dict()

    # 3. FACTURACIÓN ESTIMADA (x_facturas.proyectos) - EN DOLARES
    # Fetch project.project to map Analytic -> Project ID (Task/Project Management ID)
    domain_projects = [('analytic_account_id', 'in', project_ids)]
    projects_data = models.execute_kw(db, uid, password, 'project.project', 'search_read', [domain_projects], {'fields': ['id', 'analytic_account_id']})
    
    # Map Analytic_ID -> Project_Project_ID
    analytic_to_project_map = {p['analytic_account_id'][0]: p['id'] for p in projects_data if p['analytic_account_id']}
    all_project_ids = list(analytic_to_project_map.values())
    
    estimated_total_map = {}
    estimated_pending_map = {}
    
    if all_project_ids:
        domain_est = [('x_studio_field_sFPxe', 'in', all_project_ids)]
        # Fields: x_Monto, x_studio_field_sFPxe, x_studio_facturado (Boolean)
        fields_est = ['x_studio_field_sFPxe', 'x_Monto', 'x_studio_facturado']
        
        try:
            est_lines = models.execute_kw(db, uid, password, 'x_facturas.proyectos', 'search_read', [domain_est], {'fields': fields_est})
            df_est = pd.DataFrame(est_lines)
            
            if not df_est.empty:
                df_est['Proj_ID'] = df_est['x_studio_field_sFPxe'].apply(lambda x: x[0] if x else None)
                
                # Total Estimado (Todo)
                estimated_total_map = df_est.groupby('Proj_ID')['x_Monto'].sum().to_dict()
                
                # Pendiente (Solo False)
                # x_studio_facturado puede ser True, False, o None. Asumimos None = False (Pendiente)
                df_est['Facturado'] = df_est['x_studio_facturado'].fillna(False)
                df_pending = df_est[df_est['Facturado'] == False]
                estimated_pending_map = df_pending.groupby('Proj_ID')['x_Monto'].sum().to_dict()
                
        except Exception as e:
             st.error(f"Error fetching Estimated Billing: {e}")

    # 4. DATOS ADICIONALES: COSTO Y PROVISIONES
    # Costo: 0.533 (402), 0.511 (76), 0.517 (395), 0.531 (400), 0.521 (399)
    # Provisiones: 0.2145 (504)
    # Usamos account.analytic.line para eficiencia
    
    cost_account_ids = [76, 395, 399, 400, 402]
    provisions_account_id = 504
    
    # Cost Data
    domain_cost = [
        ('company_id', '=', 1),
        ('account_id', 'in', project_ids),
        ('general_account_id', 'in', cost_account_ids),
        ('date', '<=', end_date_str)
    ]
    
    cost_map = {}
    try:
        lines_cost = models.execute_kw(db, uid, password, 'account.analytic.line', 'search_read', [domain_cost], {'fields': ['account_id', 'amount']})
        if lines_cost:
            df_cost = pd.DataFrame(lines_cost)
            df_cost['Project_ID'] = df_cost['account_id'].apply(lambda x: x[0] if x else None)
            # Costo en analítica suele ser negativo (-). Lo mostramos positivo absoluto o sumamos tal cual?
            # User wants "Costo Total". Usually shown positive.
            # Analytic Amount = -Cost. So we sum and negate, or sum absolute?
            # Let's sum and negate.
            df_cost['Amount_Inv'] = -df_cost['amount']
            cost_map = df_cost.groupby('Project_ID')['Amount_Inv'].sum().to_dict()
    except Exception as e:
        pass

    # Provisions Data
    domain_prov = [
        ('company_id', '=', 1),
        ('account_id', 'in', project_ids),
        ('general_account_id', '=', provisions_account_id), # 504
        ('date', '<=', end_date_str)
    ]
    
    provisions_map = {}
    try:
        lines_prov = models.execute_kw(db, uid, password, 'account.analytic.line', 'search_read', [domain_prov], {'fields': ['account_id', 'amount']})
        if lines_prov:
            df_prov = pd.DataFrame(lines_prov)
            df_prov['Project_ID'] = df_prov['account_id'].apply(lambda x: x[0] if x else None)
            provisions_map = df_prov.groupby('Project_ID')['amount'].sum().to_dict()
    except Exception as e:
        pass

    # 5. INVENTARIO (STOCK ACTUAL) - x_studio_field_qCgKk
    # Link: Project -> Stock Location (via stock.location.x_studio_field_qCgKk)
    # We need to find the location where x_studio_field_qCgKk == Project ID
    
    inventory_map = {}
    if all_project_ids:
        try:
            # Find locations linked to these projects
            domain_loc = [('x_studio_field_qCgKk', 'in', all_project_ids)]
            
            # Fetch Link: Location ID -> Project ID
            locations = models.execute_kw(db, uid, password, 'stock.location', 'search_read', 
                [domain_loc], {'fields': ['id', 'x_studio_field_qCgKk', 'name']})
            
            loc_id_to_project = {}
            target_location_ids = []
            
            for l in locations:
                lid = l['id']
                # x_studio_field_qCgKk is Many2one (id, name)
                pid = l['x_studio_field_qCgKk'][0] if l['x_studio_field_qCgKk'] else None
                if pid:
                    loc_id_to_project[lid] = pid
                    target_location_ids.append(lid)
            
            if target_location_ids:
                for lid, pid in loc_id_to_project.items():
                    # Get value of Quants in this location (child_of)
                    # We use search_read because read_group often fails on 'value' 
                    # if it is a computed/non-stored field in some Odoo versions.
                    
                    try:
                        quants = models.execute_kw(db, uid, password, 'stock.quant', 'search_read',
                            [[('location_id', 'child_of', lid)]],
                            {'fields': ['value']})
                        
                        if quants:
                            val = sum(q.get('value', 0.0) for q in quants)
                            inventory_map[pid] = inventory_map.get(pid, 0.0) + val
                    except Exception as e:
                        pass
                        
        except Exception as e:
            st.error(f"Error fetching Inventory: {e}")

    # TIPO DE CAMBIO (USD -> CRC)
    # Buscamos la tasa del dólar (ID 2) para la fecha de corte
    usd_rate = 0.002 # Fallback
    try:
        # Rate closest to end_date
        domain_rate = [
            ('currency_id', '=', 2), # USD
            ('name', '<=', end_date_str),
            ('company_id', 'in', [1, False])
        ]
        rates_data = models.execute_kw(db, uid, password, 'res.currency.rate', 'search_read', 
                                       [domain_rate], 
                                       {'fields': ['rate'], 'limit': 1, 'order': 'name desc'})
        if rates_data:
            usd_rate = rates_data[0]['rate']
    except Exception as e:
        st.error(f"Error fetching Currency Rate: {e}")

    # 6. CONSOLIDACIÓN
    results = []
    
    for _, row in wip_grouped.iterrows():
        a_id = row['Project_ID']
        p_id = analytic_to_project_map.get(a_id)
        
        wip_val = row['WIP_Balance']
        is_archived = row['Is_Archived']
        
        # Facturado Real
        inc_curr = income_map_current.get(a_id, 0.0)
        inc_prev = income_map_prev.get(a_id, 0.0)
        
        # Costo y Provisiones
        cost_val = cost_map.get(a_id, 0.0)
        prov_val = provisions_map.get(a_id, 0.0)
        
        # Inventario
        inv_val = inventory_map.get(p_id, 0.0)
        
        # Estimado en USD -> CRC
        est_total_usd = estimated_total_map.get(p_id, 0.0)
        est_pending_usd = estimated_pending_map.get(p_id, 0.0)
        
        est_total_crc = 0.0
        est_pending_crc = 0.0
        
        if usd_rate > 0:
            est_total_crc = est_total_usd / usd_rate
            est_pending_crc = est_pending_usd / usd_rate
        
        results.append({
            'Proyecto': row['Project_Name'],
            'Archivado': 'Sí' if is_archived else 'No',
            'Saldo WIP (Gastos Pend.)': wip_val,
            'Costo Total': cost_val,
            'Provisiones (0.2145)': prov_val,
            'Inventario (Stock Actual)': inv_val,
            'Facturado Mes Actual': inc_curr,
            'Facturado Anterior': inc_prev,
            'Total Estimado (CRC)': est_total_crc,
            'Total Facturado': inc_curr + inc_prev,
            'Pendiente Facturar (Est.)': est_pending_crc
        })
        
    return pd.DataFrame(results)
    # Formato esperado: {'3686': 100} donde key es el ID de la cuenta analítica
    
    def extract_analytic_id(dist):
        if not dist: return None
        # Si es un diccionario (lo normal en xmlrpc para Json cols)
        if isinstance(dist, dict) and dist:
            # Retornamos el primer ID encontrado. 
            # Si hay split (varios IDs), esto tomaría uno arbitrario, pero para WIP suele ser 1 a 1.
            return list(dist.keys())[0]
        # Si Odoo devuelve string (menos probable via xmlrpc pero posible)
        return None

    df_wip['Analytic_ID_Str'] = df_wip['analytic_distribution'].apply(extract_analytic_id)
    
    # Filtrar lineas que no tienen analítica (si las hay, quedarían como 'Sin Proyecto')
    # Convertir a entero
    df_wip['Project_ID'] = pd.to_numeric(df_wip['Analytic_ID_Str'], errors='coerce')
    
    # Calcular Balance (Debit - Credit)
    # En apuntes contables: Gasto aumenta al Debe (Debit).
    # Saldo WIP = Debit - Credit.
    df_wip['WIP_Balance'] = df_wip['debit'] - df_wip['credit']
    
    # Agrupar por Project ID
    wip_grouped = df_wip.groupby('Project_ID')['WIP_Balance'].sum().reset_index()
    wip_grouped = wip_grouped[wip_grouped['WIP_Balance'] != 0]
    
    project_ids = wip_grouped['Project_ID'].dropna().unique().tolist()
    
    if not project_ids:
        return pd.DataFrame()

    # Necesitamos los Nombres de los Proyectos (Analytic Accounts)
    # Hicimos groupby solo por ID, ahora buscamos los nombres
    # project_ids son IDs de analytic.account
    try:
        analytic_accounts = models.execute_kw(db, uid, password, 'account.analytic.account', 'search_read', 
                                              [[('id', 'in', project_ids)]], {'fields': ['id', 'name']})
        aa_map = {a['id']: a['name'] for a in analytic_accounts}
    except Exception as e:
        aa_map = {}
        
    wip_grouped['Project_Name'] = wip_grouped['Project_ID'].apply(lambda x: aa_map.get(x, f"Proyecto {int(x)}"))

    # 2. FACTURACIÓN REAL (INGRESOS)
    # Mantenemos account.analytic.line para ingresos pues es más eficiente buscar por account_id (Analytic ID)
    # Buscamos ingresos asociados a estos proyectos.
    
    # Primero identificamos cuentas financieras de ingreso
    try:
        income_account_ids = models.execute_kw(db, uid, password, 'account.account', 'search', [[('account_type', 'in', ['income', 'income_other'])]])
    except:
        income_account_ids = []
        
    domain_income = [
        ('company_id', '=', 1),
        ('account_id', 'in', project_ids), # account_id en analytic_line ES la cuenta analítica
        ('date', '<=', end_date_str)
    ]
    
    if income_account_ids:
        domain_income.append(('general_account_id', 'in', income_account_ids))
    
    fields_income = ['account_id', 'amount', 'date']
    try:
        lines_income = models.execute_kw(db, uid, password, 'account.analytic.line', 'search_read', [domain_income], {'fields': fields_income})
        df_income = pd.DataFrame(lines_income)
    except:
        df_income = pd.DataFrame() # Fail gracefully
    
    # Link to Project
    income_map_current = {}
    income_map_prev = {}
    
    if not df_income.empty:
        # account_id in analytic line is relations (id, name)
        df_income['Project_ID'] = df_income['account_id'].apply(lambda x: x[0] if x else None)
        df_income['Net_Income'] = df_income['amount'] # Income is positive in analytic
        df_income['Date'] = pd.to_datetime(df_income['date']).dt.date
        
        # Split Current vs Previous
        current_mask = (df_income['Date'] >= start_of_month) & (df_income['Date'] <= end_date)
        prev_mask = (df_income['Date'] < start_of_month)
        
        income_map_current = df_income[current_mask].groupby('Project_ID')['Net_Income'].sum().to_dict()
        income_map_prev = df_income[prev_mask].groupby('Project_ID')['Net_Income'].sum().to_dict()

    # 3. FACTURACIÓN ESTIMADA (x_facturas.proyectos) - EN DOLARES
    # Fetch project.project to map Analytic -> Project ID (Task/Project Management ID)
    domain_projects = [('analytic_account_id', 'in', project_ids)]
    projects_data = models.execute_kw(db, uid, password, 'project.project', 'search_read', [domain_projects], {'fields': ['id', 'analytic_account_id']})
    
    # Map Analytic_ID -> Project_Project_ID
    analytic_to_project_map = {p['analytic_account_id'][0]: p['id'] for p in projects_data if p['analytic_account_id']}
    all_project_ids = list(analytic_to_project_map.values())
    
    estimated_total_map = {}
    estimated_pending_map = {}
    
    if all_project_ids:
        domain_est = [('x_studio_field_sFPxe', 'in', all_project_ids)]
        # Fields: x_Monto, x_studio_field_sFPxe, x_studio_facturado (Boolean)
        fields_est = ['x_studio_field_sFPxe', 'x_Monto', 'x_studio_facturado']
        
        try:
            est_lines = models.execute_kw(db, uid, password, 'x_facturas.proyectos', 'search_read', [domain_est], {'fields': fields_est})
            df_est = pd.DataFrame(est_lines)
            
            if not df_est.empty:
                df_est['Proj_ID'] = df_est['x_studio_field_sFPxe'].apply(lambda x: x[0] if x else None)
                
                # Total Estimado (Todo)
                estimated_total_map = df_est.groupby('Proj_ID')['x_Monto'].sum().to_dict()
                
                # Pendiente (Solo False)
                # x_studio_facturado puede ser True, False, o None. Asumimos None = False (Pendiente)
                df_est['Facturado'] = df_est['x_studio_facturado'].fillna(False)
                df_pending = df_est[df_est['Facturado'] == False]
                estimated_pending_map = df_pending.groupby('Proj_ID')['x_Monto'].sum().to_dict()
                
        except Exception as e:
             st.error(f"Error fetching Estimated Billing: {e}")

    # 4. DATOS ADICIONALES: COSTO Y PROVISIONES
    # Costo: 0.533 (402), 0.511 (76), 0.517 (395), 0.531 (400), 0.521 (399)
    # Provisiones: 0.2145 (504)
    # Usamos account.analytic.line para eficiencia
    
    cost_account_ids = [76, 395, 399, 400, 402]
    provisions_account_id = 504
    
    # Cost Data
    domain_cost = [
        ('company_id', '=', 1),
        ('account_id', 'in', project_ids),
        ('general_account_id', 'in', cost_account_ids),
        ('date', '<=', end_date_str)
    ]
    
    cost_map = {}
    try:
        lines_cost = models.execute_kw(db, uid, password, 'account.analytic.line', 'search_read', [domain_cost], {'fields': ['account_id', 'amount']})
        if lines_cost:
            df_cost = pd.DataFrame(lines_cost)
            df_cost['Project_ID'] = df_cost['account_id'].apply(lambda x: x[0] if x else None)
            # Costo en analítica suele ser negativo (-). Lo mostramos positivo absoluto o sumamos tal cual?
            # User wants "Costo Total". Usually shown positive.
            # Analytic Amount = -Cost. So we sum and negate, or sum absolute?
            # Let's sum and negate.
            df_cost['Amount_Inv'] = -df_cost['amount']
            cost_map = df_cost.groupby('Project_ID')['Amount_Inv'].sum().to_dict()
    except Exception as e:
        pass

    # Provisions Data
    domain_prov = [
        ('company_id', '=', 1),
        ('account_id', 'in', project_ids),
        ('general_account_id', '=', provisions_account_id) # 504
        # ('date', '<=', end_date_str) -> REMOVED
    ]
    
    provisions_map = {}
    try:
        lines_prov = models.execute_kw(db, uid, password, 'account.analytic.line', 'search_read', [domain_prov], {'fields': ['account_id', 'amount']})
        if lines_prov:
            df_prov = pd.DataFrame(lines_prov)
            df_prov['Project_ID'] = df_prov['account_id'].apply(lambda x: x[0] if x else None)
            # Provision usually credit? or debit?
            # 0.2145 sounds like Liability (Provision). 
            # If it's a provision expense (Debit), it's negative in analytic. 
            # If it's the liability account itself, analytic lines might not be generated automatically unless config exists.
            # But user asked to check if there are provisions IN account 0.2145 associated with project.
            # If entries exist, we sum them.
            # Let's just sum the raw amount for now, or invert if negative (expense view).
            # Provision is usually Liability side. Increasing provision = Credit = Positive in Accounting, but Analytic?
            # Let's assume user wants to know the Accumulated Provision Balance.
            # Analytic entries on Balance Sheet accounts are rare unless specified.
            # If 0.2145 is Liability, usually no analytic lines. 
            # If user implies the EXPENSE account that feeds provision? No, he said "cuenta 0.2145".
            # Let's stick to analytic search. If empty, maybe check move lines?
            # Validation script showed 0.2145 is "Provisión Costo Proyectos".
            # Let's try analytic first. If empty, we might need move line search like WIP.
            # However, for efficiency let's stick to analytic. If 0.2145 tracks cost per project, likely it HAS analytic.
            provisions_map = df_prov.groupby('Project_ID')['amount'].sum().to_dict()
            # If it's liability, Credits are positive ??? In analytic, usually Debit is negative, Credit is positive? 
    except Exception as e:
        pass

    # 5. INVENTARIO (STOCK ACTUAL) - x_studio_field_qCgKk
    # Link: Project -> Stock Location (via stock.location.x_studio_field_qCgKk)
    # We need to find the location where x_studio_field_qCgKk == Project ID
    
    inventory_map = {}
    if all_project_ids:
        try:
            # Find locations linked to these projects
            domain_loc = [('x_studio_field_qCgKk', 'in', all_project_ids)]
            # We need parent_path to handle children locations if logic requires, 
            # OR we can just use 'child_of' operator in search for quants.
            # Efficient way:
            # 1. Get list of locations per project.
            # 2. Search quants where location_id child_of [all_loc_ids].
            # 3. Process result to attribute to project.
            
            # Fetch Link: Location ID -> Project ID
            locations = models.execute_kw(db, uid, password, 'stock.location', 'search_read', 
                [domain_loc], {'fields': ['id', 'x_studio_field_qCgKk', 'name']})
            
            loc_id_to_project = {}
            target_location_ids = []
            
            for l in locations:
                lid = l['id']
                # x_studio_field_qCgKk is Many2one (id, name)
                pid = l['x_studio_field_qCgKk'][0] if l['x_studio_field_qCgKk'] else None
                if pid:
                    loc_id_to_project[lid] = pid
                    target_location_ids.append(lid)
            
            if target_location_ids:
                # Fetch Quants in these locations (including children)
                # We can't easily map child locations back to project without `parent_path` logic or iterative calls.
                # Iterative 'read_group' might be safer/easier if N is small.
                # Let's try read_group for each project location.
                
                # Check how many locations?
                # If < 20, loop is fine. If > 50, batching?
                # Let's try loop for correctness first. It's robust.
                
                for lid, pid in loc_id_to_project.items():
                    # Get value of Quants in this location (child_of)
                    # Domain: location_id child_of lid
                    # Field: value (sum)
                    # We use search_read because read_group often fails on 'value' 
                    # if it is a computed/non-stored field in some Odoo versions.
                    
                    try:
                        quants = models.execute_kw(db, uid, password, 'stock.quant', 'search_read',
                            [[('location_id', 'child_of', lid)]],
                            {'fields': ['value']})
                        
                        if quants:
                            val = sum(q.get('value', 0.0) for q in quants)
                            inventory_map[pid] = inventory_map.get(pid, 0.0) + val
                    except Exception as e:
                        # print(f"Error reading stock value for Loc {lid}: {e}")
                        pass
                        
        except Exception as e:
            st.error(f"Error fetching Inventory: {e}")

    # TIPO DE CAMBIO (USD -> CRC)
    # Buscamos la tasa del dólar (ID 2) para la fecha de corte
    usd_rate = 0.002 # Fallback
    try:
        # Rate closest to end_date
        domain_rate = [
            ('currency_id', '=', 2), # USD
            ('name', '<=', end_date_str),
            ('company_id', 'in', [1, False])
        ]
        rates_data = models.execute_kw(db, uid, password, 'res.currency.rate', 'search_read', 
                                       [domain_rate], 
                                       {'fields': ['rate'], 'limit': 1, 'order': 'name desc'})
        if rates_data:
            usd_rate = rates_data[0]['rate']
    except Exception as e:
        st.error(f"Error fetching Currency Rate: {e}")

    # 6. CONSOLIDACIÓN
    results = []
    
    for _, row in wip_grouped.iterrows():
        a_id = row['Project_ID']
        p_id = analytic_to_project_map.get(a_id)
        
        wip_val = row['WIP_Balance']
        
        # Facturado Real
        inc_curr = income_map_current.get(a_id, 0.0)
        inc_prev = income_map_prev.get(a_id, 0.0)
        
        # Costo y Provisiones
        cost_val = cost_map.get(a_id, 0.0)
        prov_val = provisions_map.get(a_id, 0.0)
        
        # Inventario
        inv_val = inventory_map.get(p_id, 0.0)
        
        # Estimado en USD -> CRC
        est_total_usd = estimated_total_map.get(p_id, 0.0)
        est_pending_usd = estimated_pending_map.get(p_id, 0.0)
        
        est_total_crc = 0.0
        est_pending_crc = 0.0
        
        if usd_rate > 0:
            est_total_crc = est_total_usd / usd_rate
            est_pending_crc = est_pending_usd / usd_rate
        
        results.append({
            'Proyecto': row['Project_Name'],
            'Saldo WIP (Gastos Pend.)': wip_val,
            'Costo Total': cost_val,
            'Provisiones (0.2145)': prov_val,
            'Inventario (Stock Actual)': inv_val,
            'Facturado Mes Actual': inc_curr,
            'Facturado Anterior': inc_prev,
            # 'Total Estimado (CRC)': est_total_crc,
            'Total Facturado': inc_curr + inc_prev,
            'Pendiente Facturar (Est.)': est_pending_crc
        })
        
    return pd.DataFrame(results)

def vista_wip_report():
    st.title("🚧 Reporte WIP (Proyectos en Proceso)")
    st.markdown("Análisis de Cuenta 0.11531 vs Facturación, Costos, Provisiones e Inventario.")
    st.divider()
    
    col1, col2 = st.columns([1, 3])
    
    with col1:
        st.subheader("Configuración")
        cut_date = st.date_input("Fecha de Corte (Fin de Mes)", datetime.date.today())
        st.caption("Nota: El Inventario mostrado es el Stock Actual (no histórico). WIP muestra solo partidas no conciliadas.")
        btn = st.button("Generar Reporte WIP", type="primary")
        
    with col2:
        if btn:
            with st.spinner('Procesando datos contables, proyectos e inventario...'):
                uid, models, db, pwd = get_odoo_connection()
                if uid:
                    df = fetch_wip_data(uid, models, db, pwd, cut_date)
                    
                    if not df.empty:
                        # Ordenar por Saldo WIP descendente
                        df = df.sort_values(by='Saldo WIP (Gastos Pend.)', ascending=False)
                        
                        # Totales Generales
                        t_wip = df['Saldo WIP (Gastos Pend.)'].sum()
                        t_fact_mes = df['Facturado Mes Actual'].sum()
                        t_est_pende = df['Pendiente Facturar (Est.)'].sum()
                        t_inv = df['Inventario (Stock Actual)'].sum()
                        
                        st.markdown(f"##### Estado Global al {cut_date}")
                        c1, c2, c3, c4 = st.columns(4)
                        
                        # Helper for smaller metric display
                        def small_metric(col, label, value):
                            col.markdown(
                                f"""
                                <div style="border-left: 3px solid #f0f2f6; padding-left: 10px;">
                                    <p style="font-size: 12px; color: #555; margin-bottom: 0px;">{label}</p>
                                    <p style="font-size: 18px; font-weight: 600; color: #FFF; margin-top: 0px;">{value}</p>
                                </div>
                                """, 
                                unsafe_allow_html=True
                            )

                        small_metric(c1, "Total en WIP", f"₡ {t_wip:,.2f}")
                        small_metric(c2, "Facturado Este Mes", f"₡ {t_fact_mes:,.2f}")
                        small_metric(c3, "Est. Pendiente Facturar", f"₡ {t_est_pende:,.2f}")
                        small_metric(c4, "Inventario Actual", f"₡ {t_inv:,.2f}")
                        st.divider()
                        
                        # DEFINICIÓN DE FORMATOS
                        format_dict = {
                            'Saldo WIP (Gastos Pend.)': '₡ {:,.2f}',
                            'Costo Total': '₡ {:,.2f}',
                            'Provisiones (0.2145)': '₡ {:,.2f}',
                            'Inventario (Stock Actual)': '₡ {:,.2f}',
                            'Facturado Mes Actual': '₡ {:,.2f}',
                            'Facturado Anterior': '₡ {:,.2f}',
                            # 'Total Estimado (CRC)': '₡ {:,.2f}', 
                            'Total Facturado': '₡ {:,.2f}',
                            'Pendiente Facturar (Est.)': '₡ {:,.2f}' 
                        }

                        # SPLIT LOGIC
                        # 1. WIP para Reclasificar (Tienen facturación en el mes actual)
                        df_reclass = df[df['Facturado Mes Actual'] != 0]
                        df_review = df[df['Facturado Mes Actual'] == 0]
                        
                        # --- SECCIÓN 1: RECLASIFICAR ---
                        st.subheader("🔄 WIP para Reclasificar (con Facturación este mes)")
                        if not df_reclass.empty:
                            st.dataframe(df_reclass.style.format(format_dict), use_container_width=True)
                            st.caption(f"Total WIP Reclasificar: ₡ {df_reclass['Saldo WIP (Gastos Pend.)'].sum():,.2f}")
                        else:
                            st.info("No hay proyectos con facturación este mes.")
                            
                        st.divider()
                        
                        # --- SECCIÓN 2: REVISAR ---
                        st.subheader("📋 WIP para Revisar (Sin facturación reciente)")
                        if not df_review.empty:
                            st.dataframe(df_review.style.format(format_dict), use_container_width=True)
                            st.caption(f"Total WIP Revisar: ₡ {df_review['Saldo WIP (Gastos Pend.)'].sum():,.2f}")
                        else:
                            st.info("No hay proyectos pendientes de revisión.")

                        # Excel Export (All Data)
                        output = io.BytesIO()
                        with pd.ExcelWriter(output, engine='openpyxl') as writer:
                            df.to_excel(writer, index=False, sheet_name='WIP_Completo')
                            if not df_reclass.empty:
                                df_reclass.to_excel(writer, index=False, sheet_name='Para_Reclasificar')
                            if not df_review.empty:
                                df_review.to_excel(writer, index=False, sheet_name='Para_Revisar')
                            
                        st.download_button(
                            "📥 Descargar Excel Completo (WIP)", 
                            output.getvalue(), 
                            f"WIP_Report_{cut_date}.xlsx", 
                            "application/vnd.ms-excel"
                        )
                        
                    else:
                        st.info("No se encontraron saldos en la cuenta WIP para la fecha seleccionada.")

def main():
    st.sidebar.title("Menú")
    opciones = {
        "Inicio": vista_inicio, 
        "Antigüedad de Saldos": vista_reporte,
        "Ventas Retail": vista_ventas_retail,
        "Reporte WIP": vista_wip_report
    }
    selection = st.sidebar.radio("Ir a:", list(opciones.keys()))
    opciones[selection]()

if __name__ == "__main__":
    main()


